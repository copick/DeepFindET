# This script is adapted from a public GitHub repository.
# Original source: https://github.com/deep-finder/cryoet-deepfinder/tree/master
# Author: Inria,  Emmanuel Moebel, Charles Kervrann
# License: GPL v3.0

from contextlib import redirect_stdout # for writing txt file
from openpyxl import load_workbook # for excel files
from openpyxl import Workbook
from copy import deepcopy
from lxml import etree
import numpy as np
import os

import warnings
warnings.simplefilter('ignore') # to mute some warnings produced when opening the tomos with mrcfile

def add_obj(objlIN, label, coord, obj_id=None, tomo_idx=None, orient=(None,None,None), cluster_size=None):
    obj = {
        'tomo_idx': tomo_idx,
        'obj_id'  : obj_id  ,
        'label'   : label   ,
        'x'       :coord[2] ,
        'y'       :coord[1] ,
        'z'       :coord[0] ,
        'psi'     :orient[0],
        'phi'     :orient[1],
        'the'     :orient[2],
        'cluster_size':cluster_size
    }
    # return objlIN.append(obj)
    objlIN.append(obj)
    return objlIN

def disp(objlIN):
    """Prints objl in terminal"""
    for p in range(len(objlIN)):
        tidx  = objlIN[p]['tomo_idx']
        objid = objlIN[p]['obj_id']
        lbl   = objlIN[p]['label']
        x     = objlIN[p]['x']
        y     = objlIN[p]['y']
        z     = objlIN[p]['z']
        psi   = objlIN[p]['psi']
        phi   = objlIN[p]['phi']
        the   = objlIN[p]['the']
        csize = objlIN[p]['cluster_size']

        strout = 'obj ' + str(p) + ': ('
        if tidx!=None:
            strout = strout + 'tomo_idx:' + str(tidx) + ', '
        if objid!=None:
            strout = strout + 'obj_id:' + str(objid) + ', '
        strout = strout + 'lbl:' + str(lbl) + ', x:' + str(x) + ', y:' + str(y) + ', z:' + str(z) + ', '
        if psi!=None or phi!=None or the!=None:
            strout = strout + 'psi:' + str(psi) + ', phi:' + str(phi) + ', the:' + str(the) + ', '
        if csize!=None:
            strout = strout + 'cluster_size:' + str(csize)
        strout = strout + ')'

        print(strout)


def read(filename):
    """Reads object list. Handles .xml and .xlsx files, according to what extension the file has.

    Args:
        filename (str): '/path/to/file.ext' with '.ext' either '.xml' or '.xlsx'

    Returns:
        list of dict
    """
    data_format = os.path.splitext(filename)
    if data_format[1] == '.xml':
        objl = read_xml(filename)
    elif data_format[1] == '.xlsx':
        objl = read_excel(filename)
    else:
        print('/!\ DeepFinder can only read object lists in .xml and .xlsx formats')
    return objl


def write(objl, filename):
    """Writes object list. Can write .xml and .xlsx files, according to the extension specified in filename.

    Args:
        objl (list of dict)
        filename (str): '/path/to/file.ext' with '.ext' either '.xml' or '.xlsx'
    """
    data_format = os.path.splitext(filename)
    if data_format[1] == '.xml':
        write_xml(objl, filename)
    elif data_format[1] == '.xlsx':
        write_excel(objl, filename)
    else:
        print('/!\ DeepFinder can only write object lists in .xml and .xlsx formats')


def read_xml(filename):
    tree = etree.parse(filename)
    objl_xml = tree.getroot()

    objlOUT = []
    for p in range(len(objl_xml)):
        tidx  = objl_xml[p].get('tomo_idx')
        objid = objl_xml[p].get('obj_id')
        lbl   = objl_xml[p].get('class_label')
        x     = objl_xml[p].get('x')
        y     = objl_xml[p].get('y')
        z     = objl_xml[p].get('z')
        psi   = objl_xml[p].get('psi')
        phi   = objl_xml[p].get('phi')
        the   = objl_xml[p].get('the')
        csize = objl_xml[p].get('cluster_size')

        # if facultative attributes exist, then cast to correct type:
        if tidx!=None:
            tidx = int(tidx)
        if objid!=None:
            objid = int(objid)
        if csize!=None:
            csize = int(csize)
        if psi!=None or phi!=None or the!=None:
            psi = float(psi)
            phi = float(phi)
            the = float(the)

        add_obj(objlOUT, tomo_idx=tidx, obj_id=objid, label=int(lbl), coord=(float(z), float(y), float(x)), orient=(psi,phi,the), cluster_size=csize)
    return objlOUT

def write_xml(objlIN, filename):
    objl_xml = etree.Element('objlist')
    for idx in range(len(objlIN)):
        tidx  = objlIN[idx]['tomo_idx']
        objid = objlIN[idx]['obj_id']
        lbl   = objlIN[idx]['label']
        x     = objlIN[idx]['x']
        y     = objlIN[idx]['y']
        z     = objlIN[idx]['z']
        psi   = objlIN[idx]['psi']
        phi   = objlIN[idx]['phi']
        the   = objlIN[idx]['the']
        csize = objlIN[idx]['cluster_size']

        obj = etree.SubElement(objl_xml, 'object')
        if tidx!=None:
            obj.set('tomo_idx', str(tidx))
        if objid!=None:
            obj.set('obj_id', str(objid))
        obj.set('class_label' , str(lbl))
        obj.set('x'           , '%.3f' % x)
        obj.set('y'           , '%.3f' % y)
        obj.set('z'           , '%.3f' % z)
        if psi!=None:
            obj.set('psi', '%.3f' % psi)
        if phi!=None:
            obj.set('phi', '%.3f' % phi)
        if the!=None:
            obj.set('the', '%.3f' % the)
        if csize!=None:
            obj.set('cluster_size', str(csize))

    tree = etree.ElementTree(objl_xml)
    tree.write(filename, pretty_print=True)

def read_excel(filename):
    wb = load_workbook(filename, enumerate)
    sheet = wb.worksheets[0]
    rows = sheet.max_row

    objl = []
    for idx in range(2,rows+1): # in excel, '0' is for col prop and '1' is col titles
        tidx  = sheet['A'+str(idx)].value
        objid = sheet['B'+str(idx)].value
        lbl   = sheet['C'+str(idx)].value
        x     = sheet['D'+str(idx)].value
        y     = sheet['E'+str(idx)].value
        z     = sheet['F'+str(idx)].value
        psi   = sheet['G'+str(idx)].value
        phi   = sheet['H'+str(idx)].value
        the   = sheet['I'+str(idx)].value
        csize = sheet['J'+str(idx)].value

        # if facultative attributes exist, then cast to correct type:
        if tidx != None:
            tidx = int(tidx)
        if objid != None:
            objid = int(objid)
        if csize != None:
            csize = int(csize)
        if psi != None or phi != None or the != None:
            psi = float(psi)
            phi = float(phi)
            the = float(the)

        add_obj(objl, tomo_idx=tidx, obj_id=objid, label=int(lbl), coord=(float(z), float(y), float(x)),
                orient=(psi, phi, the), cluster_size=csize)
    return objl

def write_excel(objl, filename):
    wb = Workbook()
    #sheet = wb.create_sheet(title='Object list')
    sheet = wb.active
    sheet.title = "Object list"
    sheet['A1'] = 'Tomo IDX' # col titles
    sheet['B1'] = 'Object ID'
    sheet['C1'] = 'Class label'
    sheet['D1'] = 'x'
    sheet['E1'] = 'y'
    sheet['F1'] = 'z'
    sheet['G1'] = 'psi'
    sheet['H1'] = 'phi'
    sheet['I1'] = 'theta'
    sheet['J1'] = 'Cluster size'

    for idx in range(len(objl)):
        sheet['A'+str(idx+2)] = objl[idx]['tomo_idx']
        sheet['B'+str(idx+2)] = objl[idx]['obj_id']
        sheet['C'+str(idx+2)] = objl[idx]['label']
        sheet['D'+str(idx+2)] = objl[idx]['x']
        sheet['E'+str(idx+2)] = objl[idx]['y']
        sheet['F'+str(idx+2)] = objl[idx]['z']
        sheet['G'+str(idx+2)] = objl[idx]['psi']
        sheet['H'+str(idx+2)] = objl[idx]['phi']
        sheet['I'+str(idx+2)] = objl[idx]['the']
        sheet['J'+str(idx+2)] = objl[idx]['cluster_size']

    wb.save(filename=filename)

# label can be int or str (is casted to str)
def get_class(objlIN, label):
    """
    Get all objects of specified class.

    Args:
        objl (list of dict)
        label (int)
    Returns:
        list of dict: contains only objects from class 'label'
    """
    idx_class = []
    for idx in range(len(objlIN)):
        if str(objlIN[idx]['label'])==str(label):
            idx_class.append(idx)

    objlOUT = []
    for idx in range(len(idx_class)):
        objlOUT.append(objlIN[idx_class[idx]])
    return objlOUT

def above_thr(objlIN, thr):
    """
    Args:
        objl (list of dict)
        thr (float): threshold

    Returns:
        list of dict: contains only objects with cluster size >= thr
    """
    idx_thr = []
    for idx in range(len(objlIN)):
        csize = objlIN[idx]['cluster_size']
        if csize != None:
            if csize>=thr:
                idx_thr.append(idx)
        else:
            print('/!\ Object ' + str(idx) + ' has no attribute cluster_size')

    objlOUT = []
    for idx in range(len(idx_thr)):
        objlOUT.append( objlIN[idx_thr[idx]] )
    return objlOUT

def above_thr_per_class(objlIN, lbl_list, thr_list):
    objlOUT = []
    for lbl in lbl_list:
        objl_class = get_class(objlIN, lbl)
        objl_class = above_thr(objl_class, thr_list[lbl-1])
        for p in range(len(objl_class)):
            objlOUT.append(objl_class[p])
    return objlOUT